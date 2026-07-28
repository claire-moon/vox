#!/usr/bin/env python3
# SPDX-License-Identifier: GPL-3.0-or-later
"""Build the deterministic VOX + DIGS QA feedback workbook."""

from __future__ import annotations

import argparse
import csv
import os
import re
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE = ROOT / "qa" / "VOX_QA_CHECKPOINTS.csv"
DEFAULT_OUTPUT = ROOT / "qa" / "VOX_QA_FEEDBACK.xlsx"
FIXED_TIME = datetime(2026, 1, 1, 0, 0, 0)
ZIP_TIME = (2026, 1, 1, 0, 0, 0)
CORE_TIMESTAMP = b"2026-01-01T00:00:00Z"

HEADER_FILL = PatternFill("solid", fgColor="2B2118")
HEADER_FONT = Font(color="FFF3D6", bold=True)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
BLOCKED_FILL = PatternFill("solid", fgColor="FCE4D6")
SEVERE_FILL = PatternFill("solid", fgColor="F4B084")
RESULTS = '"Not Run,Pass,Fail,Blocked,Skip"'
SEVERITIES = '"NOTE,LOW,MEDIUM,HIGH,BLOCKER"'
ISSUE_STATUSES = '"Open,Confirmed,Needs Info,Fixed,Retest,Closed"'


def arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate qa/VOX_QA_FEEDBACK.xlsx from the checkpoint CSV"
    )
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--check",
        action="store_true",
        help="verify that an existing output matches a fresh deterministic build",
    )
    return parser.parse_args()


def read_checkpoints(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    required = {
        "ID",
        "Area",
        "Title",
        "Setup",
        "Steps",
        "Expected Result",
        "Evidence Requested",
    }
    if not rows:
        raise ValueError(f"no checkpoints found in {path}")
    if set(rows[0]) != required:
        raise ValueError(f"checkpoint columns must be {sorted(required)}")
    ids = [row["ID"] for row in rows]
    if any(not item for item in ids) or len(ids) != len(set(ids)):
        raise ValueError("checkpoint IDs must be nonempty and unique")
    return rows


def style_header(sheet, columns: int) -> None:
    for cell in sheet[1][:columns]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.protection = Protection(locked=True)
    sheet.row_dimensions[1].height = 32
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False


def set_widths(sheet, widths: Sequence[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width


def add_list_validation(sheet, formula: str, cells: str, prompt: str) -> None:
    validation = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        error="Choose a value from the list.",
        errorTitle="Invalid QA value",
        prompt=prompt,
        promptTitle="VOX QA",
    )
    validation.errorStyle = "stop"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(cells)


def add_checkpoint_sheet(workbook: Workbook, rows: Iterable[dict[str, str]]) -> None:
    sheet = workbook.active
    sheet.title = "Checkpoints"
    sheet.sheet_properties.tabColor = "C55A11"
    headers = [
        "ID",
        "Area",
        "Title",
        "Setup",
        "Steps",
        "Expected Result",
        "Evidence Requested",
        "Result",
        "Severity if Failed",
        "Build ID",
        "Tester",
        "Tested At UTC",
        "Evidence Path",
        "Notes",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                row["ID"],
                row["Area"],
                row["Title"],
                row["Setup"],
                row["Steps"],
                row["Expected Result"],
                row["Evidence Requested"],
                "Not Run",
                "NOTE",
                "",
                "",
                "",
                "",
                "",
            ]
        )
    style_header(sheet, len(headers))
    set_widths(sheet, [15, 15, 27, 38, 52, 52, 36, 13, 18, 18, 18, 21, 34, 42])
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for cell in row[7:14]:
            cell.fill = INPUT_FILL
    sheet.auto_filter.ref = f"A1:N{sheet.max_row}"
    add_list_validation(sheet, RESULTS, f"H2:H{sheet.max_row}", "Record the checkpoint outcome.")
    add_list_validation(
        sheet,
        SEVERITIES,
        f"I2:I{sheet.max_row}",
        "Choose impact only when the result is Fail or Blocked.",
    )
    sheet.conditional_formatting.add(
        f"H2:H{sheet.max_row}", FormulaRule(formula=["H2=\"Pass\""], fill=PASS_FILL)
    )
    sheet.conditional_formatting.add(
        f"H2:H{sheet.max_row}", FormulaRule(formula=["H2=\"Fail\""], fill=FAIL_FILL)
    )
    sheet.conditional_formatting.add(
        f"H2:H{sheet.max_row}", FormulaRule(formula=["H2=\"Blocked\""], fill=BLOCKED_FILL)
    )


def add_issue_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Issues")
    sheet.sheet_properties.tabColor = "C00000"
    headers = [
        "Issue ID",
        "Checkpoint ID",
        "Summary",
        "Severity",
        "Status",
        "Build ID",
        "Platform",
        "Reproduction Steps",
        "Expected",
        "Actual",
        "Evidence Path",
        "Owner",
        "GitHub URL",
        "Reported At UTC",
        "Notes",
    ]
    sheet.append(headers)
    for _ in range(50):
        sheet.append([""] * len(headers))
    style_header(sheet, len(headers))
    set_widths(sheet, [16, 16, 34, 14, 15, 18, 24, 52, 40, 40, 34, 18, 38, 21, 40])
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.fill = INPUT_FILL
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.auto_filter.ref = f"A1:O{sheet.max_row}"
    add_list_validation(sheet, SEVERITIES, f"D2:D{sheet.max_row}", "Rate player and test impact.")
    add_list_validation(sheet, ISSUE_STATUSES, f"E2:E{sheet.max_row}", "Track the issue lifecycle.")
    sheet.conditional_formatting.add(
        f"D2:D{sheet.max_row}", FormulaRule(formula=["OR(D2=\"BLOCKER\",D2=\"HIGH\")"], fill=SEVERE_FILL)
    )


def add_environment_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Environment")
    sheet.sheet_properties.tabColor = "548235"
    sheet.append(["Field", "Value", "Source", "Notes"])
    fields = [
        ("Tester alias", "Manual", "Use a public alias; do not enter an email address."),
        ("Build ID", "Bundle or release", "Record the exact bundle or release identifier."),
        ("VOX version", "Bundle or release", "Expected demo version is v0.0.3."),
        ("Git commit", "Bundle manifest", "Use the public commit hash when supplied."),
        ("Binary SHA-256", "Cockpit or checksum file", "Identifies the exact tested executable."),
        ("Operating system", "System settings", "Name and release only."),
        ("Distribution", "/etc/os-release or equivalent", "Linux distribution when applicable."),
        ("Kernel", "uname -r", "Do not include the host name."),
        ("Architecture", "uname -m", "Examples: x86_64 or aarch64."),
        ("CPU model", "System settings", "Include model; omit hardware serial numbers."),
        ("Logical CPU count", "System settings", "Record the available logical CPUs."),
        ("Memory", "System settings", "Installed or available RAM."),
        ("GPU model", "System settings", "Presence does not imply VOX GPU acceleration."),
        ("GPU driver", "System settings", "Record driver and version if available."),
        ("SDL2 version", "Package manager", "Runtime or linked SDL2 version."),
        ("Display server", "System settings", "Examples: X11 Wayland Quartz or Win32."),
        ("Desktop", "System settings", "Desktop environment or window manager."),
        ("Display resolution / refresh", "System settings", "Record the tested display mode."),
        ("Test lane", "Manual", "Quick Full Automation Build-only or another clearly named lane."),
        ("QA run ID", "Cockpit or manual", "Use the cockpit run ID when available."),
        ("Frame caps tested", "VOX Options", "List 15 30 60 90 120 144 and Unlimited as tested."),
        ("Cap qualification result", "VOX startup / self-test", "List supported and visibly gated caps; do not infer support."),
        ("Lightfield tiers tested", "VOX Options", "List Compatibility Balanced and Showcase."),
        ("Laptop Mode", "VOX Options", "Record Off or On for every parity/hash run."),
        ("Dummy Mode", "VOX Options", "Record Off or On for bark-cooldown evidence."),
        ("Input devices", "Manual", "Keyboard and pointing device model or type."),
        ("Controller model / SDL name", "SDL / manual", "General model/name only; omit unique serial data."),
        ("Controller SDL GUID", "SDL diagnostics", "Controller mapping identity; never substitute a hardware serial."),
        ("Controller transport", "Manual", "Record USB Bluetooth or other without unique serial data."),
        ("Controller prompt family", "VOX UI", "Record Nintendo Xbox PlayStation or generic."),
        ("Controller mapping path", "VOX diagnostics", "Record SDL mapped or raw joystick fallback."),
        ("P1 / P2 input ownership", "VOX Options", "Record AUTO KEYBOARD CONTROLLER and claimed pad per local player."),
        ("Aim calibration", "VOX Options", "Record sensitivity deadzone aim slowdown and whether calibration ran."),
        ("P1 / P2 rope mode", "VOX Options", "Record Hold or Toggle independently for each local player."),
        ("Haptics level / availability", "VOX Options / SDL", "Record Off Low Normal Heavy and available blocked or unsupported."),
        ("Haptic mixer self-test result", "VOX automation log", "Record off low normal heavy near and far values; this does not prove physical rumble."),
        ("Audio device", "System settings", "Use a general device label; omit unique IDs."),
        ("Audio backend / cadence result", "SDL / VOX self-test", "Record backend and callback underrun or cadence result when exposed."),
        ("Power state", "Manual", "AC or battery and relevant performance profile."),
        ("Map style / landform / seed", "Match setup", "Required for terrain collision debris and camera reproduction."),
        ("Local players / bots", "Match setup", "Record the active slot composition."),
        ("FX profile", "Match setup", "Record Retro Standard or Carnage for simulation/hash comparison."),
        ("Deterministic load self-test result", "VOX automation log", "Record the 600-tick slots activity awake cells canonical hash and exit status; no timing threshold applies."),
        ("Named-bench performance qualification", "VOX automation log", "Record average p95 maximum activity hash and power profile only from the named i7-10750H laptop bench."),
        ("Initial / final state hash", "F1 / automation log", "Record both when comparing caps or Laptop Mode."),
        ("Frame hash / smoke SHA-256", "Automation log", "Identifies deterministic presentation evidence."),
        ("xleak version", "xleak --version", "Cockpit dependency version."),
        ("Test started UTC", "Manual", "Use ISO 8601: YYYY-MM-DDTHH:MM:SSZ."),
        ("Test ended UTC", "Manual", "Use ISO 8601: YYYY-MM-DDTHH:MM:SSZ."),
        ("Evidence sharing consent", "Manual", "Enter Yes only after reviewing the privacy checklist."),
    ]
    for field, source, notes in fields:
        sheet.append([field, "", source, notes])
    style_header(sheet, 4)
    set_widths(sheet, [29, 42, 31, 56])
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        row[1].fill = INPUT_FILL
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.auto_filter.ref = f"A1:D{sheet.max_row}"
    add_list_validation(
        sheet,
        '"Yes,No"',
        f"B{sheet.max_row}",
        "Confirm only after reviewing qa/README.md.",
    )


def build_workbook(source: Path, output: Path) -> None:
    rows = read_checkpoints(source)
    workbook = Workbook()
    workbook.properties.creator = "VOX contributors"
    workbook.properties.lastModifiedBy = "VOX deterministic workbook generator"
    workbook.properties.title = "VOX + DIGS v0.0.3 QA Feedback"
    workbook.properties.subject = "Portable demo acceptance and issue evidence"
    workbook.properties.description = "Generated from qa/VOX_QA_CHECKPOINTS.csv"
    workbook.properties.created = FIXED_TIME
    workbook.properties.modified = FIXED_TIME
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.calculation.calcMode = "manual"
    add_checkpoint_sheet(workbook, rows)
    add_issue_sheet(workbook)
    add_environment_sheet(workbook)

    output.parent.mkdir(parents=True, exist_ok=True)
    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=".vox-qa-", suffix=".xlsx", dir=output.parent
    )
    os.close(file_descriptor)
    temporary = Path(temporary_name)
    try:
        workbook.save(temporary)
        normalized = temporary.with_suffix(".normalized.xlsx")
        with zipfile.ZipFile(temporary, "r") as source_zip:
            with zipfile.ZipFile(
                normalized, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
            ) as output_zip:
                for name in sorted(source_zip.namelist()):
                    original = source_zip.getinfo(name)
                    data = source_zip.read(name)
                    if name == "docProps/core.xml":
                        data = re.sub(
                            rb"(<dcterms:(?:created|modified)\b[^>]*>)[^<]*(</dcterms:(?:created|modified)>)",
                            lambda match: match.group(1)
                            + CORE_TIMESTAMP
                            + match.group(2),
                            data,
                        )
                    info = zipfile.ZipInfo(name, ZIP_TIME)
                    info.compress_type = zipfile.ZIP_DEFLATED
                    info.external_attr = original.external_attr
                    info.create_system = 0
                    output_zip.writestr(info, data)
        os.replace(normalized, output)
    finally:
        temporary.unlink(missing_ok=True)
        temporary.with_suffix(".normalized.xlsx").unlink(missing_ok=True)


def validate_workbook(path: Path, checkpoint_count: int) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    if workbook.sheetnames != ["Checkpoints", "Issues", "Environment"]:
        raise ValueError(f"unexpected sheets: {workbook.sheetnames}")
    if workbook["Checkpoints"].max_row != checkpoint_count + 1:
        raise ValueError("checkpoint workbook row count does not match the CSV")
    for sheet in workbook.worksheets:
        if str(sheet.freeze_panes) != "A2" or not sheet.auto_filter.ref:
            raise ValueError(f"{sheet.title} is missing frozen panes or filters")
    workbook.close()


def main() -> int:
    args = arguments()
    source = args.source.resolve()
    output = args.output.resolve()
    rows = read_checkpoints(source)
    if args.check:
        if not output.is_file():
            raise SystemExit(f"workbook does not exist: {output}")
        with tempfile.TemporaryDirectory(prefix="vox-qa-check-") as directory:
            candidate = Path(directory) / output.name
            build_workbook(source, candidate)
            if output.read_bytes() != candidate.read_bytes():
                raise SystemExit(
                    f"{output} is stale; run tools/build-qa-workbook.py"
                )
        validate_workbook(output, len(rows))
        print(f"QA workbook is current: {output}")
        return 0
    build_workbook(source, output)
    validate_workbook(output, len(rows))
    print(f"Wrote {output} with {len(rows)} checkpoints")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
